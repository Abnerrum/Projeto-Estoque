import io
from datetime import datetime
from flask import Blueprint, render_template, send_file, request
from flask_login import login_required
from app import db
from app.models.produto import Produto
from app.models.movimentacao import Movimentacao

rel_bp = Blueprint("relatorios", __name__, url_prefix="/relatorios")


@rel_bp.route("/")
@login_required
def index():
    # Produtos críticos (abaixo do mínimo)
    todos = Produto.query.filter_by(ativo=True).all()
    criticos = [p for p in todos if p.abaixo_minimo]

    # Totais
    total_produtos = len(todos)
    total_criticos = len(criticos)

    # Últimas movimentações
    ultimas = Movimentacao.query.order_by(Movimentacao.criado_em.desc()).limit(10).all()

    # Entradas e saídas do mês atual
    agora = datetime.utcnow()
    entradas_mes = db.session.query(db.func.coalesce(db.func.sum(Movimentacao.quantidade), 0)).filter(
        Movimentacao.tipo == "entrada",
        db.func.strftime("%Y-%m", Movimentacao.criado_em) == agora.strftime("%Y-%m")
    ).scalar()

    saidas_mes = db.session.query(db.func.coalesce(db.func.sum(Movimentacao.quantidade), 0)).filter(
        Movimentacao.tipo == "saida",
        db.func.strftime("%Y-%m", Movimentacao.criado_em) == agora.strftime("%Y-%m")
    ).scalar()

    return render_template("relatorios/index.html",
                           criticos=criticos,
                           total_produtos=total_produtos,
                           total_criticos=total_criticos,
                           ultimas=ultimas,
                           entradas_mes=int(entradas_mes),
                           saidas_mes=int(saidas_mes))


@rel_bp.route("/exportar-excel")
@login_required
def exportar_excel():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return "openpyxl não instalado.", 500

    wb = Workbook()

    # Aba 1: Posição de estoque
    ws1 = wb.active
    ws1.title = "Posição de Estoque"
    cabecalho = ["Código", "Produto", "Categoria", "Unidade", "Saldo Atual", "Estoque Mínimo", "Status"]
    ws1.append(cabecalho)
    for cell in ws1[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="1D4ED8")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    for p in Produto.query.filter_by(ativo=True).order_by(Produto.nome).all():
        status = "CRÍTICO" if p.abaixo_minimo else "OK"
        cat = p.categoria.nome if p.categoria else "-"
        ws1.append([p.codigo, p.nome, cat, p.unidade, p.saldo_atual, p.estoque_minimo, status])

    # Aba 2: Histórico de movimentações
    ws2 = wb.create_sheet("Movimentações")
    ws2.append(["Data", "Tipo", "Produto", "Código", "Quantidade", "Usuário", "Justificativa"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)

    for m in Movimentacao.query.order_by(Movimentacao.criado_em.desc()).all():
        ws2.append([
            m.criado_em.strftime("%d/%m/%Y %H:%M"),
            m.tipo.upper(),
            m.produto.nome,
            m.produto.codigo,
            m.quantidade,
            m.usuario.nome,
            m.justificativa or "-",
        ])

    # Ajusta largura das colunas automaticamente
    for ws in [ws1, ws2]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    nome = f"estoque_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(output, as_attachment=True, download_name=nome,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
